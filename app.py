import logging
import traceback
from flask import Flask, render_template, request, redirect, session, jsonify, send_file
import json
from azure.storage.blob import BlobServiceClient
import openpyxl,io
from azure.cosmosdb.table.tableservice import TableService
from azure.cosmosdb.table.models import Entity
import requests
from azure.common import AzureMissingResourceHttpError
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os
from io import BytesIO
from email.mime.image import MIMEImage
app = Flask(__name__)
app.secret_key = 'sanketh2341562'
# Replace with your actual Azure Storage account credentials
storage_account_name = 'blobdatabase234'
storage_account_key = 'xnu1+DDZO1s9n2y4qYU6J39WyBHZVMLIk6pWVl4bAi8WuvPrOJM9WuTppAAPAEWU3liXUnUm9NFx+AStzG1QAw=='
container_name = 'login'
excel_file_name = 'credentials.xlsx'
excel_file_name1 = 'studentform.xlsx'

def save_student_data_to_excel(name, roll_no, email):
    # Replace with your Azure Table storage account connection string
    connection_string = "DefaultEndpointsProtocol=https;AccountName=blobdatabase234;AccountKey=xnu1+DDZO1s9n2y4qYU6J39WyBHZVMLIk6pWVl4bAi8WuvPrOJM9WuTppAAPAEWU3liXUnUm9NFx+AStzG1QAw==;EndpointSuffix=core.windows.net"
    table_service = TableService(connection_string=connection_string)

    # Define the name of your Azure Table
    table_name = "studentdata1"

    # Create the table if it doesn't exist
    if not table_service.exists(table_name):
        table_service.create_table(table_name)

    # Create a new entity and set its properties (columns)
    student_entity = {
        "PartitionKey": name,
        "RowKey": roll_no,
        "Email": email,
    }

    # Insert the entity into the Azure Table
    table_service.insert_or_replace_entity(table_name, student_entity)

# Your existing functions and routes here...

def save_student_data_to_excel1(name, roll_no, email, status):
    # Replace with your Azure Table storage account connection string
    connection_string = "DefaultEndpointsProtocol=https;AccountName=blobdatabase234;AccountKey=xnu1+DDZO1s9n2y4qYU6J39WyBHZVMLIk6pWVl4bAi8WuvPrOJM9WuTppAAPAEWU3liXUnUm9NFx+AStzG1QAw==;EndpointSuffix=core.windows.net"
    table_service = TableService(connection_string=connection_string)

    # Define the name of your Azure Table
    table_name = "Records"

    # Create the table if it doesn't exist
    if not table_service.exists(table_name):
        table_service.create_table(table_name)

    # Create a new entity and set its properties (columns)
    student_entity = {
        "PartitionKey": name,
        "RowKey": roll_no,
        "Email": email,
        "Status": status,
    }
    # Insert the entity into the Azure Table
    table_service.insert_or_replace_entity(table_name, student_entity)


@app.route('/delete_all_rows', methods=['POST'])
def delete_all_rows():
    # Connect to your Azure Table storage
    table_service = TableService(connection_string='DefaultEndpointsProtocol=https;AccountName=blobdatabase234;AccountKey=xnu1+DDZO1s9n2y4qYU6J39WyBHZVMLIk6pWVl4bAi8WuvPrOJM9WuTppAAPAEWU3liXUnUm9NFx+AStzG1QAw==;EndpointSuffix=core.windows.net')

    # Get all entities from the table
    entities = table_service.query_entities('studentdata1')

    # Delete each entity (row) from the table
    for entity in entities:
        table_service.delete_entity('studentdata1', entity.PartitionKey, entity.RowKey)

    return render_template('form.html')



@app.route('/', methods=['GET', 'POST'])
def login():
    error_message = None  # Initialize error_message as None
    if request.method == 'POST':
        # Retrieve login form data
        username = request.form['username']
        password = request.form['password']

        # Connect to Azure Storage Blob
        blob_service_client = BlobServiceClient(account_url=f"https://{storage_account_name}.blob.core.windows.net", credential=storage_account_key)
        container_client = blob_service_client.get_container_client(container_name)

        # Download the Excel file from Azure Blob Storage
        blob_client = container_client.get_blob_client(excel_file_name)
        with open(excel_file_name, "wb") as f:
            f.write(blob_client.download_blob().readall())

        # Authenticate user from the Excel file
        authenticated = authenticate_user(username, password, excel_file_name)

        if authenticated:
            return redirect('/homepage')
        else:
            error_message = "Please check your credentials, sir/madam please type username as frt and password as frt(This is for demo purpose only)"  # Set error message

    return render_template('login.html', error_message=error_message)  # Pass error_message to template


@app.route('/homepage', methods=['GET', 'POST'])
def HomePage():
    # You can add logic here to fetch data or prepare content for the homepage
    # For example, fetching user-specific data, etc.
    return render_template('HomePage.html')

@app.route('/logout')
def logout():
    # Add your logout logic here, for example, clearing the session
    session.clear()
    return redirect('/')

def authenticate_user(username, password, excel_file):
    # Open the Excel file
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    # Iterate through rows and compare credentials
    for row in sheet.iter_rows(values_only=True):
        if row[0] == username and row[1] == password:
            return True
    return False

@app.route('/form')
def student_form():
    return render_template('form.html')


@app.route('/submit_form', methods=['GET', 'POST'])
def submit_form():
    if request.method == 'POST':
        # Retrieve form data from the request
        name = request.form['name']
        roll_no = request.form['roll_no']
        email = request.form['email']

        # Save the data to the Excel file in Azure Blob Storage
        save_student_data_to_excel(name, roll_no, email)

        # Set the 'submitted' variable to True
        submitted = True

        return render_template('form.html',submitted=submitted)
    
#Start of email code
#
#
# Function to fetch student data and QR code URLs from Azure Table storage
def fetch_student_data_from_table():
    # Replace with your Azure Table storage account connection string
    connection_string = "DefaultEndpointsProtocol=https;AccountName=blobdatabase234;AccountKey=xnu1+DDZO1s9n2y4qYU6J39WyBHZVMLIk6pWVl4bAi8WuvPrOJM9WuTppAAPAEWU3liXUnUm9NFx+AStzG1QAw==;EndpointSuffix=core.windows.net"
    table_service = TableService(connection_string=connection_string)

    # Replace with your Azure Table name
    table_name = "studentdata1"

    # Fetch all entities from the Azure Table
    entities = table_service.query_entities(table_name)

    # Initialize a list to store student data and QR code URLs
    qr_codes_data = []

    # Iterate through each entity and get the student data and QR code URLs
    # Iterate through each entity and get the student data and QR code URLs
    for entity in entities:
        name = entity.get('PartitionKey', '')
        roll_no = entity.get('RowKey', '')
        email = entity.get('Email', '')
        qr_codes_data.append({"name": name, "roll_no": roll_no, "email": email})

    return qr_codes_data
# Function to send an email with an attachment
# Flask route to generate QR codes, send emails, and display success message

# Flask route to fetch QR codes and display success message
@app.route('/generate_qr_codes', methods=['POST'])
def generate_qr_codes():
    # Make a GET request to the Azure Function endpoint that generates QR codes
    response = requests.get("https://qrcode111.azurewebsites.net/api/generateallqrcodes")  # Replace with the URL of your Azure Function
    # Check the status code to determine if the request was successful
    if response.status_code == 200:
        # Render the form.html template with the success message
        return render_template('form.html')
    else:
        # Handle the case where the request was not successful
        # You can return an error message or redirect to an error page
        return render_template('form.html')

def send_email(smtp_username, smtp_password, sender_email, receiver_email, subject, message, qr_code_urls):
    # Set up the MIMEText object with the email content
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject
    msg.attach(MIMEText(message, 'plain'))

    for qr_code_url in qr_code_urls:
        # Fetch the QR code image data from the Blob storage URL
        response = requests.get(qr_code_url)
        if response.status_code == 200:
            qr_code_data = response.content
            qr_code_filename = os.path.basename(qr_code_url)

            # Attach the QR code image to the email
            qr_code_attachment = MIMEBase('image', 'png')  # Specify the correct MIME type and subtype here
            qr_code_attachment.set_payload(qr_code_data)
            encoders.encode_base64(qr_code_attachment)
            qr_code_attachment.add_header('Content-Disposition', f'attachment; filename={qr_code_filename}')
            msg.attach(qr_code_attachment)

    try:
        # Connect to the SMTP server and send the email
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(smtp_username, smtp_password)
            server.send_message(msg)

        logging.info(f"Email sent successfully to {receiver_email}")
        return True
    except Exception as e:
        logging.error(f"Failed to send email to {receiver_email}. Error: {e}")
        logging.error(traceback.format_exc())  # Print the full traceback for detailed error analysis
        return False
#####

qr_codes_directory = os.path.join(os.getcwd(), "qrcodes")
attachment_directory = os.path.join(os.getcwd(), "attachments")

def download_qr_code_images(container_client, qr_codes_data):
    qr_code_images = []
    for student_info in qr_codes_data:
        name, roll_no = student_info['name'], student_info['roll_no']
        qr_code_filename = f"{name}_{roll_no}.png"
        local_file_path = os.path.join(qr_codes_directory, qr_code_filename)
        with open(local_file_path, "wb") as local_file:
            local_file.write(container_client.get_blob_client(qr_code_filename).download_blob().readall())
        qr_code_images.append(local_file_path)

    return qr_code_images

@app.route('/sendmail', methods=['POST'])
def sendmail():
    # Replace with your SMTP email credentials
    smtp_username = 'spalaksha@gmail.com'
    smtp_password = 'hicqrmqyaiffmtwq'
    sender_email = 'spalaksha@gmail.com'

    # Fetch student data and QR code URLs from Azure Table storage
    qr_codes_data = fetch_student_data_from_table()

    # Fetch the Blob Service Client
    blob_service_client = BlobServiceClient.from_connection_string("DefaultEndpointsProtocol=https;AccountName=blobdatabase234;AccountKey=xnu1+DDZO1s9n2y4qYU6J39WyBHZVMLIk6pWVl4bAi8WuvPrOJM9WuTppAAPAEWU3liXUnUm9NFx+AStzG1QAw==;EndpointSuffix=core.windows.net")
    container_client = blob_service_client.get_container_client("qrcodes")

    # Loop through the QR codes data and send emails to each student
    for student_info in qr_codes_data:
        name = student_info['name']
        roll_no = student_info['roll_no']
        email = student_info['email']
        message = f"Dear {name},\n\nPlease download this QR code to receive attendance for tomorrow's Annual Day event.\n\nBest Regards,\nYour friend"

        # Get the URL of the QR code image from Azure Blob Storage
        qr_code_url = container_client.get_blob_client(f"qrcodes/{name}_{roll_no}.png").url

        # Send the email with the QR code URL as an attachment
        send_email(smtp_username, smtp_password, sender_email, email, "Attendance QR Code", message, [qr_code_url])

    # Render the form.html template with the success message
    return render_template('form.html', message="Generated QR Codes Successfully! Check your email for the QR codes.")

#End of email


@app.route('/scanner', methods=['POST'])
def handle_scanner_data():
    if request.method == 'POST':
        qr_data = json.loads(request.data)
        name = qr_data.get('name', '')
        roll_no = qr_data.get('roll_no', '')
        email = qr_data.get('email', '')
        status = "Present"

        # Save the data to the Excel file
        save_student_data_to_excel1(name, roll_no, email, status)

        # Send a response back to the frontend
        response_data = {"status": "success", "message": "QR code data processed successfully!"}
        return jsonify(response_data)

    # Handle other HTTP methods (if needed)
    return jsonify({"status": "error", "message": "Invalid request method."}), 405

# Function to fetch student data from Azure Table and save it to Excel

def save_student_data_to_excel2():
    # Replace with your Azure Table storage account connection string
    connection_string = "DefaultEndpointsProtocol=https;AccountName=blobdatabase234;AccountKey=xnu1+DDZO1s9n2y4qYU6J39WyBHZVMLIk6pWVl4bAi8WuvPrOJM9WuTppAAPAEWU3liXUnUm9NFx+AStzG1QAw==;EndpointSuffix=core.windows.net"
    table_service = TableService(connection_string=connection_string)

    # Define the name of your Azure Table
    table_name = "Records"

    # Query all entities from the Azure Table
    entities = table_service.query_entities(table_name)

    # Create an Excel workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the headers
    headers = ["Name", "Roll No", "Email", "Status"]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)

    # Write the student data to the worksheet
    row_num = 2
    for student in entities:
        student_data = [student.PartitionKey, student.RowKey, student.Email, student.Status]
        for col_num, data in enumerate(student_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=data)
        row_num += 1

    # Save the Excel workbook to a BytesIO buffer
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    # Return the Excel data as a BytesIO object
    return excel_buffer

# Flask route to handle the download of the Records table
@app.route('/download_records')
def download_records():
    # Get the Excel data from the "Records" table in Azure
    excel_data = save_student_data_to_excel2()

    # Send the Excel data as a downloadable file to the user
    return send_file(
        excel_data,
        as_attachment=True,
        attachment_filename='Records.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Home Page
@app.route('/')
def home_page():
    return render_template('HomePage.html')

# Video Stream Page
@app.route('/video_stream')
def video_stream_page():
    return render_template('video_stream.html')


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000)